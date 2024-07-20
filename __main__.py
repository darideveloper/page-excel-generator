import os
import shutil
import zipfile
from time import sleep
from io import BytesIO

import openpyxl
import requests
from PIL import Image

from dotenv import load_dotenv


# Env variables
load_dotenv()
DOMAIN = os.getenv("DOMAIN")
IMAGES_FOLDER = os.getenv("IMAGES_FOLDER")
EXCEL_SHEET = os.getenv("EXCEL_SHEET")


class PageGenerator():
    def __init__(self):
        """ Initialize PageGenerator object
        
        """
        
        print("Initializing PageGenerator...")
        
        # Paths
        self.current_folder = os.path.dirname(os.path.abspath(__file__))
        self.template_path = os.path.join(self.current_folder, "template.html")
        self.excel_path = os.path.join(self.current_folder, "input.xlsx")
        self.htmls_folder = os.path.join(self.current_folder, "htmls")
        
        # Create folders if not exists
        os.makedirs(self.htmls_folder, exist_ok=True)
        
        # Data variables
        self.excel_data = []
        self.excel_header = []
        
        # Validation data
        self.columns = {
            "row": 1,
            "names": [
                "url",
                # title required in all pages
                "title",
                "description",
                "image url",
                "site name"
            ]
        }
        self.columns_row = self.columns["row"]
        self.columns_names = self.columns["names"]
        
        # Load excel data and validate columns
        self.__load_excel_data__()
        self.__save_header__()
        self.__validate_excel_columns__()
        
        # Process data
        self.__replace_images_paths__()
        
        # Process folders
        self.__clean_htmls_folder__()
        
    def __load_excel_data__(self):
        """ Read excel data and save in instance
        """
        
        print("Loading excel data...")
        
        # Validate excel path
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        # Read excel
        wb = openpyxl.load_workbook(self.excel_path)
        current_sheet = wb[EXCEL_SHEET]
        
        rows = current_sheet.max_row
        columns = current_sheet.max_column

        data = []
        for row in range(1, rows + 1):

            row_data = []
            for column in range(1, columns + 1):
                cell_data = current_sheet.cell(row, column).value
                row_data.append(cell_data)

            data.append(row_data)

        self.excel_data = data
        
    def __save_header__(self):
        """ Save in instance the excel header """
        
        template_row = self.columns_row
        excel_header = self.excel_data[template_row - 1]
        self.excel_header = excel_header
        
    def __validate_excel_columns__(self):
        """ Check specific columns and column's order in excel """
        
        print("Validating excel columns...")
        
        # Validete excel header
        for column_name in self.columns_names:
            if column_name not in self.excel_header:
                error = f"Column '{column_name}' not found in excel"
                error += f"\nExcel columns: {self.excel_header}"
                raise ValueError(error)
            
    def __replace_images_paths__(self):
        """ Replace images paths using relative paths and images folder """
        
        # Identify images columns
        images_columns_indexes = []
        for column_name in self.excel_header:
            if "image" in str(column_name):
                column_index = self.excel_header.index(column_name)
                images_columns_indexes.append(column_index)
        
        # Replace images paths
        for row in self.excel_data[self.columns_row:]:
            for column_index in images_columns_indexes:
                image_file = row[column_index]
                new_image_path = f"{DOMAIN}/{IMAGES_FOLDER}/{image_file}"
                row[column_index] = new_image_path

    def __clean_htmls_folder__(self):
        """ Clean html folder """
        
        print("Cleaning html folder...")
        
        for file in os.listdir(self.htmls_folder):
            file_path = os.path.join(self.htmls_folder, file)
            if os.path.isfile(file_path):
                os.remove(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
    
    def __get_image_size__(self, image_url: str) -> tuple[int, int]:
        """ Return image with and height

        Args:
            image_url (str): image to download
            
        Returns:
            tuple[int, int]: image width and height
        """
        
        try:
            image_res = requests.get(image_url)
            image_res.raise_for_status()
        except Exception:
            return 0, 0
        image = Image.open(BytesIO(image_res.content))
        return image.size
           
    def generate_pages(self):
        """ Generate pages using template with and excel data """
        
        print("Generating pages...")
        
        template_content = open(self.template_path, "r").read()
        
        # generate each html file with excel data
        for row in self.excel_data[self.columns_row:]:
            
            sleep(0.1)
            
            # Create folder
            title_index = self.excel_header.index("description")
            page_title = row[title_index]
            if not page_title:
                continue
            slug = page_title.lower().replace(" ", "-")
            
            print(f"\tGenerating page: {slug}")
            
            # Create html folder
            html_folder = os.path.join(self.htmls_folder, slug)
            os.makedirs(html_folder, exist_ok=True)
            html_path = os.path.join(html_folder, "index.html")
            
            # Replace each cell in template
            content = template_content
            for cell in row:
                cell_index = row.index(cell)
                current_column_name = self.excel_header[cell_index]
                
                content = content.replace(f"[{current_column_name}]", cell)
                
            # Remplace image size
            image_url = row[self.excel_header.index("image url")]
            image_width, image_height = self.__get_image_size__(image_url)
            if not image_width and not image_height:
                print(f"\t\tError getting image size: {image_url}")
                continue
            content = content.replace("[image width]", str(image_width))
            content = content.replace("[image height]", str(image_height))
            
            # Save html file with content
            with open(html_path, "w", encoding="utf-8") as file:
                file.write(content)
                
    def compress_htmls(self):
        """ Compress htmls folders to zip """
        
        print("Compressing htmls...")
        
        output_path = os.path.join(self.htmls_folder, "pages.zip")
        folders = os.listdir(self.htmls_folder)
        folders = [os.path.join(self.htmls_folder, folder) for folder in folders]
        
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            
            for folder in folders:
                
                # Walk the directory tree
                for root, dirs, files in os.walk(folder):
                    for file in files:
                        # Create the relative path for the file and write it to the zip
                        file_path = os.path.join(root, file)
                        relative_path = os.path.relpath(
                            file_path,
                            os.path.join(folder, '..')
                        )
                        zipf.write(file_path, relative_path)
    
    
if __name__ == "__main__":
    pg = PageGenerator()
    pg.generate_pages()
    pg.compress_htmls()